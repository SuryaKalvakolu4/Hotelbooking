import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from zipfile import BadZipFile

class HotelBookingApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Hotel Booking System")
        self.master.geometry("400x350")

        self.load_data()

        self.user_label = tk.Label(master, text="User Name:")
        self.user_label.pack()
        self.user_entry = tk.Entry(master)
        self.user_entry.pack()

        self.room_label = tk.Label(master, text="Room Number:")
        self.room_label.pack()
        self.room_entry = tk.Entry(master)
        self.room_entry.pack()

        self.book_button = tk.Button(master, text="Book", command=self.book_room)
        self.book_button.pack()

        self.view_button = tk.Button(master, text="View Bookings", command=self.view_bookings)
        self.view_button.pack()

        self.delete_button = tk.Button(master, text="Delete Booking", command=self.delete_booking)
        self.delete_button.pack()

    def load_data(self):
        try:
            self.wb = load_workbook("C:\\Users\\surya\\Desktop\\Hotelbooking\\bookings.xlsx")
        except BadZipFile:
            # Handle the case of an empty or corrupted file
            self.wb = Workbook()
            self.wb.save("C:\\Users\\surya\\Desktop\\Hotelbooking\\bookings.xlsx")
            self.wb = load_workbook("C:\\Users\\surya\\Desktop\\Hotelbooking\\bookings.xlsx")
        self.sheet = self.wb.active

    def book_room(self):
        user_name = self.user_entry.get().strip()
        room_number = self.room_entry.get().strip()

        if not user_name or not room_number:
            messagebox.showerror("Error", "Please enter both user name and room number!")
            return

        if not room_number.isdigit() or int(room_number) <= 0 or int(room_number) > 100:
            messagebox.showerror("Error", "Invalid room number!")
            return

        for row in self.sheet.iter_rows(values_only=True):
            if row[0] == int(room_number):
                messagebox.showerror("Error", "Room already booked!")
                return

        self.sheet.append((int(room_number), user_name))
        self.wb.save("C:\\Users\\surya\\Desktop\\Hotelbooking\\bookings.xlsx")
        messagebox.showinfo("Success", "Room booked successfully!")

    def view_bookings(self):
        bookings = "\n".join([f"Room {row[0]} booked by {row[1]}" for row in self.sheet.iter_rows(values_only=True)])
        if not bookings:
            bookings = "No bookings yet."
        messagebox.showinfo("Bookings", bookings)

    def delete_booking(self):
        room_number = self.room_entry.get().strip()
        if not room_number.isdigit() or int(room_number) <= 0 or int(room_number) > 100:
            messagebox.showerror("Error", "Invalid room number!")
            return

        found = False
        for row in self.sheet.iter_rows():
            if row[0].value == int(room_number):
                self.sheet.delete_rows(row[0].row)
                self.wb.save("C:\\Users\\surya\\Desktop\\Hotelbooking\\bookings.xlsx")
                messagebox.showinfo("Success", f"Booking for Room {room_number} deleted successfully!")
                found = True
                break

        if not found:
            messagebox.showerror("Error", "Booking not found!")

def main():
    root = tk.Tk()
    HotelBookingApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
